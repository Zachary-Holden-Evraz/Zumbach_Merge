# Zumbach Merge
# Note that this was created in Python and made into a windows executable file

import os, sys
from sys import exit
import shutil
import openpyxl as pyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import *
from tkinter import simpledialog
from tkinter.filedialog import askopenfilename
import fnmatch
import pandas as pd
from threading import Thread

# Ask the User to give the files - this will allow this to work on any Windows computer
main_file = askopenfilename(title = 'File you want to update (Zumbach Database)') # Excel File we want to add all the data into
dir_with_files = simpledialog.askstring(title = 'Folder path prompt', prompt = "Please type in the full path of the folder containing your files:    ")

def Merge_the_Steel():
    # Gives the names of the individual pieces
    piece_names = []
    for file in os.listdir(dir_with_files):
        try:
            piece_name = file[6:file.rindex('.01')]
            piece_names.append(piece_name)
        except:
            pass

    # Creates Subfolders for each piece
    for name in piece_names:
        try:
            path = os.path.join(dir_with_files, str(name))
            os.mkdir(path)
        except:
            pass

    if stop == 0:
        text= Label(app, text = "Stopped")
        text.grid()
        sys.exit()
        
    # This will separate the program by piece names and move them into the appropriate subfolder
    for name in piece_names:
        for file in os.listdir(dir_with_files):
            os.chdir(dir_with_files)
            if file.startswith('CHART'):
                try:
                    piece_name = file[6:file.rindex('.0')]
                except:
                    pass
                if fnmatch.fnmatch(piece_name, name):
                    # We have to resave the csv files as xlsx files.  We will delete them after
                    # Start with the numbered files
                    if fnmatch.fnmatch(file, '*01.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                    elif fnmatch.fnmatch(file, '*02.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                    elif fnmatch.fnmatch(file, '*03.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                    elif fnmatch.fnmatch(file, '*04.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                    elif fnmatch.fnmatch(file, '*05.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                    elif fnmatch.fnmatch(file, '*06.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                    elif fnmatch.fnmatch(file, '*07.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
            if file.startswith('PIECE'):
                try: 
                    piece_name = file[6:file.rindex('.CSV')]
                except:
                    pass 
                if fnmatch.fnmatch(piece_name, name):
                    # Non-numbered files
                    if fnmatch.fnmatch(file, '*AM.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                    elif fnmatch.fnmatch(file, '*PM.csv'):
                        target_dir = dir_with_files + '\\' + piece_name
                        shutil.move(file, target_dir)
                if fnmatch.fnmatch(file, 'PIECE.csv'):
                    old_file = pd.read_csv(file)
                    old_file.to_excel('Red.xlsx', index = False)
    if stop == 0:
        text= Label(app, text = "Stopped")
        text.grid()
        sys.exit()

    # Need to relist the piece names since the files are no longer in the parent dir.  
    piece_names = []
    for name in os.listdir(dir_with_files):
        path = os.path.join(dir_with_files, name)
        if os.path.isdir(path):
            piece_names.append(name)
        if fnmatch.fnmatch(name,'PIECE.csv'):
            os.chdir(dir_with_files)
            old_file = pd.read_csv(name)
            old_file.to_excel('Red.xlsx', index = False)
        else:
            pass
    if stop == 0:
        text= Label(app, text = "Stopped")
        text.grid()
        sys.exit()

    data_files = [] # Empty array for deleting files later
    # We need to set up Red_data early since it is in a different directory (parent) 
    try:
        os.chdir(dir_with_files)
        for file in os.listdir(dir_with_files):
            if fnmatch.fnmatch(file, 'Red.xlsx'):
                Red_data = file
    except:
        pass
    red = pyxl.load_workbook(Red_data)
    red_sheet = red.worksheets[0]

    # This will start getting the data per piece
    for name in piece_names:
        data_files = [] # This array will make it faster to delete the excess files later
        folder = os.path.join(dir_with_files, name)
        for file in os.listdir(folder):
            os.chdir(folder)
            if file.startswith('CHART'):
                try:
                    piece_name = file[6:file.rindex('.0')]
                except:
                    pass
                if fnmatch.fnmatch(piece_name, name):
                    # We have to resave the csv files as xlsx files.  We will delete them after
                    # Start with the numbered files
                    if fnmatch.fnmatch(file, '*01.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Blue.xlsx', index = False)
                        data_files.append(file)
                    elif fnmatch.fnmatch(file, '*02.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Light_green.xlsx', index = False)
                        data_files.append(file)
                    elif fnmatch.fnmatch(file, '*03.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Beige.xlsx', index = False)
                        data_files.append(file)
                    elif fnmatch.fnmatch(file, '*04.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Grey.xlsx', index = False)
                        data_files.append(file)
                    elif fnmatch.fnmatch(file, '*05.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Orange.xlsx', index = False)
                        data_files.append(file)
                    elif fnmatch.fnmatch(file, '*06.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Green.xlsx', index = False)
                        data_files.append(file)
                    elif fnmatch.fnmatch(file, '*07.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Blueish.xlsx', index = False)
                        data_files.append(file)
            if file.startswith('PIECE'):
                try: 
                    piece_name = file[6:file.rindex('.CSV')]
                except:
                    pass 
                if fnmatch.fnmatch(piece_name, name):
                    # Non-numbered files
                    if fnmatch.fnmatch(file, '*AM.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Yellow.xlsx', index = False)
                        data_files.append(file)
                        yellow = True
                    elif fnmatch.fnmatch(file, '*PM.csv'):
                        old_file = pd.read_csv(file)
                        old_file.to_excel('Yellow.xlsx', index = False)
                        data_files.append(file)
                        yellow = True
                if fnmatch.fnmatch(file, 'PIECE.csv'):
                    old_file = pd.read_csv(file)
                    old_file.to_excel('Red.xlsx', index = False)
        if stop == 0:
            text= Label(app, text = "Stopped")
            text.grid()
            sys.exit()

        try:
            if yellow == True:
                pass
            else:
                Yellow_path = askopenfilename(title = 'First Iteration of the piece (Yellow file)')
                Yellow_data = pd.read_csv(Yellow_path).to_excel('Yellow.xlsx', index = False)
        except:
            Yellow_path = askopenfilename(title = 'First Iteration of the piece (Yellow file)')
            Yellow_data = pd.read_csv(Yellow_path).to_excel('Yellow.xlsx', index = False)

            # Not sure about brown yet.  Most pieces do not seem to have both yellow and brown
            # Brown_path = askopenfilename(title = 'Second Iteration of the piece (Brown file)')
            # if Brown_path != '':
            #     Brown_data = pd.read_csv(Brown_path).to_excel('Brown.xlsx', index = False)


        # We have to relist the file to read it properly
        for file in os.listdir(folder):
            os.chdir(folder)
            if fnmatch.fnmatch(file, 'Blue.xlsx'):
                Blue_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Light_green.xlsx'):
                Light_green_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Beige.xlsx'):
                Beige_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Grey.xlsx'):
                Grey_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Orange.xlsx'):
                Orange_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Green.xlsx'):
                Green_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Blueish.xlsx'):
                Blueish_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Red.xlsx'):
                Red_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Yellow.xlsx'):
                Yellow_data = file
                data_files.append(file)
            if fnmatch.fnmatch(file, 'Brown.xlsx'):
                Brown_data = file
                data_files.append(file)


        # Setting up the main file everything will go into
        main_data = pyxl.load_workbook(main_file)
        main_sheet = main_data.worksheets[0]

        # Setting up the active workbook
        workbook = Workbook()
        worksheet = workbook.active
        row_count = main_sheet.max_row
        column_count = main_sheet.max_column
        if stop == 0:
            text= Label(app, text = "Stopped")
            text.grid()
            sys.exit()

        # Make sure data in the main file stays (because this will overwite existing file)
        for i in range (1, row_count + 1):
            for j in range (1, column_count + 1):
                # reading cell value from source file
                c = main_sheet.cell(row = i, column = j)
                # writing the read value to destination file
                worksheet.cell(row = i, column = j).value = c.value


        # Reading all the Data from each file
        try:
            red = pyxl.load_workbook(Red_data)
            red_sheet = red.worksheets[0]
            print('Red is being Reloaded')
        except:
            pass
        yellow = pyxl.load_workbook(Yellow_data)
        yellow_sheet = yellow.worksheets[0]
        try: # Not all pieces have a brown file
            brown = pyxl.load_workbook(Brown_data)
            brown_sheet = brown.worksheets[0]
        except:
            brown = ''
        orange = pyxl.load_workbook(Orange_data)
        orange_sheet = orange.worksheets[0]
        green = pyxl.load_workbook(Green_data)
        green_sheet = green.worksheets[0]
        blue = pyxl.load_workbook(Blue_data)
        blue_sheet = blue.worksheets[0]
        try: # Not all pieces have a 2.csv (light_green)
            light_green = pyxl.load_workbook(Light_green_data)
            light_green_sheet = light_green.worksheets[0]
        except:
            light_green = ''
        beige = pyxl.load_workbook(Beige_data)
        beige_sheet = beige.worksheets[0]
        blueish = pyxl.load_workbook(Blueish_data)
        blueish_sheet = blueish.worksheets[0]
        grey = pyxl.load_workbook(Grey_data)
        grey_sheet = grey.worksheets[0]
        if stop == 0:
            text= Label(app, text = "Stopped")
            text.grid()
            sys.exit()


        # Move data from red_data into the main workbook
            # Product column: D - 4th column # Other columns to copy: G:J - 7:10
            # Write to columns B:F - 2:6
        # Product value
        c = red_sheet.cell(row = 2, column = 4) 
        worksheet.cell(row = row_count + 1, column = 2).value = c.value
        # Other values from the red data (piece)
        col_num = 3 # index to continue column number on main sheet
        for i in range(7, 10+1):
            c = red_sheet.cell(row = 2, column = i)
            # will go into columns C:F - 3:6
            worksheet.cell(row = row_count+1, column = col_num).value = c.value
            col_num += 1    

        # Move data from yellow_data into the main workbook
        for i in range(2, 8+1): # Rows to copy: 2:8
            for j in range(3, 11+1): # Columns to copy: C:K - 3:11
                c = yellow_sheet.cell(row = i, column = j)
                if c.value == None: continue # If there is noting in the column, skip it
                else: pass
                worksheet.cell(row = row_count+1, column = col_num).value = c.value
                col_num += 1

        # Move data from brown_data into the main workbook (if it exists)
        if brown != '':
            for i in range(2, 8+1): # Rows to copy: 2:8
                for j in range(8, 11+1): # Columns to copy: H:K - 8:11
                    c = brown_sheet.cell(row = i, column = j)
                    if c.value == None: continue # If there is nothing in the column, skip it
                    else: pass
                    worksheet.cell(row = row_count+1, column = col_num).value = c.value
                    col_num += 1
        else:
            col_num = 95 # This is in case the brown data doesn't exist in this piece

        # Move data from other files into every 3-4 columns in the main workbook
        max_row = orange_sheet.max_row
        max_col = orange_sheet.max_column
        for row_to_copy in range(2, max_row + 1):
            for j in range(1, 4+1): # This is the only file that we will copy the first column from; the 1st column should be the same in every file
                c = orange_sheet.cell(row = row_to_copy, column = j)
                worksheet.cell(row = row_count+1, column = col_num).value = c.value
                col_num +=1
            for j in range(2, 4+1):
                c = green_sheet.cell(row = row_to_copy, column = j)
                worksheet.cell(row = row_count+1, column = col_num).value = c.value
                col_num +=1
            for j in range(2, 4+1):
                c = blue_sheet.cell(row = row_to_copy, column = j)
                worksheet.cell(row = row_count+1, column = col_num).value = c.value
                col_num +=1
            for j in range(2, 4+1):
                if light_green != '':
                    c = light_green_sheet.cell(row = row_to_copy, column = j)
                    worksheet.cell(row = row_count+1, column = col_num).value = c.value
                    col_num +=1
                else:
                    col_num += 1
            for j in range(2, 4+1):
                c = beige_sheet.cell(row = row_to_copy, column = j)
                worksheet.cell(row = row_count+1, column = col_num).value = c.value
                col_num +=1
            for j in range(2, 4+1):
                c = blueish_sheet.cell(row = row_to_copy, column = j)
                worksheet.cell(row = row_count+1, column = col_num).value = c.value
                col_num +=1
            for j in range(2, 4+1):
                c = grey_sheet.cell(row = row_to_copy, column = j)
                worksheet.cell(row = row_count+1, column = col_num).value = c.value
                col_num +=1
        if stop == 0:
            text= Label(app, text = "Stopped")
            text.grid()
            sys.exit()

        # The next parts of this are because the workbook formatting gets reset during the above processes

        # Merge and color cells for neatness
        # Red
        worksheet.merge_cells('B1:F1') # merge the cells
        cell = worksheet.cell(row = 1, column = 2)
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center') # center the cell
        cell.fill = PatternFill("solid", start_color="FF0000") # fill the cell with a color
        # Yellow
        worksheet.merge_cells('G1:BN1')
        cell = worksheet.cell(row = 1, column = 7)
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        cell.fill = PatternFill("solid", start_color="FFFF00")
        # Brown
        worksheet.merge_cells('BO1:CP1')
        cell = worksheet.cell(row = 1, column = 67)
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        cell.fill = PatternFill("solid", start_color="FFCC99")
        # Orange
        i = 95
        while i < column_count + 1:
            if worksheet.cell(row = 1, column = i).value == None:
                worksheet.cell(row = 1, column = i).value = '5'
            worksheet.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+3)
            cell = worksheet.cell(row = 1, column = i)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            cell.fill = PatternFill("solid", start_color="FFCC00")
            i += 22
        # Green
        i = 99 # starting column number 
        while i < column_count + 1:
            if worksheet.cell(row = 1, column = i).value == None:
                worksheet.cell(row = 1, column = i).value = '6'
            worksheet.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+2)
            cell = worksheet.cell(row = 1, column = i)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            cell.fill = PatternFill("solid", start_color="99CC00") 
            i += 22
        # Blue
        i = 102
        while i < column_count + 1:
            if worksheet.cell(row = 1, column = i).value == None:
                worksheet.cell(row = 1, column = i).value = '1'
            worksheet.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+2)
            cell = worksheet.cell(row = 1, column = i)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            cell.fill = PatternFill("solid", start_color="00CCFF")
            i += 22
        # Light green
        i = 105
        while i < column_count + 1:
            if worksheet.cell(row = 1, column = i).value == None:
                worksheet.cell(row = 1, column = i).value = '2'
            worksheet.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+2)
            cell = worksheet.cell(row = 1, column = i)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            cell.fill = PatternFill("solid", start_color="339966")
            i += 22
        # Beige
        i = 108
        while i < column_count + 1:
            if worksheet.cell(row = 1, column = i).value == None:
                worksheet.cell(row = 1, column = i).value = '3'
            worksheet.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+2)
            cell = worksheet.cell(row = 1, column = i)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            cell.fill = PatternFill("solid", start_color="FFFF99")
            i += 22
        # Blueish
        i = 111
        while i < column_count + 1:
            if worksheet.cell(row = 1, column = i).value == None:
                worksheet.cell(row = 1, column = i).value = '7'
            worksheet.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+2)
            cell = worksheet.cell(row = 1, column = i)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            cell.fill = PatternFill("solid", start_color="33CCCC")
            i += 22
        # Grey
        i = 114
        while i < column_count + 1:
            if worksheet.cell(row = 1, column = i).value == None:
                worksheet.cell(row = 1, column = i).value = '4'
            worksheet.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+2)
            cell = worksheet.cell(row = 1, column = i)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            cell.fill = PatternFill("solid", start_color="C0C0C0")
            i += 22

        # Apply Borders to every cell in the 1st and 2nd rows
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        for i in range(1,2+1):
            for j in range(2, column_count):
                worksheet.cell(row = i, column = j).border = thin_border

        # Change the widths of the columns to increase readability 
        def as_text(value):
            if value is None:
                return ""
            return str(value)
        ### Does not get max properly, so I added 2 to the length.  Not perfect, but better
        for column_cells in worksheet.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells) + 2
            worksheet.column_dimensions[pyxl.utils.get_column_letter(column_cells[0].column)].width = length

        # Must delete data from Red file in case the program is restarted, as to not copy previously used data
        red_sheet.delete_rows(2)
        os.chdir(dir_with_files)
        red.save(filename = Red_data)

        # Save the excel file and close the workbooks
        workbook.save(filename = main_file)
        os.chdir(dir_with_files)
        shutil.copy(main_file, 'Zumbach Database Backup.xlsx')
        workbook.close()
        # Not sure if this part is necessary, must test and see if it saves time
        main_data.close()
        # try:
        #     red.close()
        # except:
        #     pass
        yellow.close()
        try:
            brown.close()
        except:
            pass
        orange.close()
        green.close()
        blue.close()
        try:
            light_green.close()
        except:
            pass
        beige.close()
        blueish.close()
        grey.close()

        # Delete the extra xslx files we created earlier.
        for file in data_files:
            os.chdir(folder)
            if fnmatch.fnmatch(file, file):
                # print(file)
                os.remove(file)
        
        # Remove the folder for the piece that is finished
        os.chdir(dir_with_files)
        path = os.path.join(dir_with_files, name)
        os.rmdir(path)

        if stop == 0:
            text= Label(app, text = "Stopped")
            text.grid()
            sys.exit()


# This will start the program in its own thread so that the stop button does not freeze
def start_thread():
    # Assign global variable and initialize value
    global stop
    stop = 1

    # Create and launch a thread 
    t = Thread (target = Merge_the_Steel)
    t.start()

def stop():
    # Assign global variable and set value to stop
    global stop
    stop = 0

root_win = tk.Tk()
root_win.title("Zumbach Merge")
root_win.geometry('100x100')
app = Frame(root_win)
app.grid()
start_button = Button(app, text="Start the Merge",command=start_thread)
stop_button = Button(app, text="Stop Merging",command=stop)

start_button.grid()
stop_button.grid()

app.mainloop()
